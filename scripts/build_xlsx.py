#!/usr/bin/env python3
"""Assemble the organised lat/lon dataset + rotation results into an Excel workbook."""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCREEN = [14443, 14444, 14445, 14446, 14447, 14448, 14449, 14452, 14453, 14454, 14455]
daily = pd.read_csv("ar_daily_positions.csv")
rot   = pd.read_csv("ar_rotation_rates.csv")

daily["date"] = pd.to_datetime(daily["date"], utc=True).dt.strftime("%Y-%m-%d")
daily = daily.rename(columns={
    "ar_noaanum": "NOAA", "date": "날짜(UTC)",
    "stony_lon": "Stonyhurst 경도(°)", "lat": "위도(°)",
    "carr_lon": "Carrington 경도(°)", "carr_lat": "Carrington 위도(°)",
    "n": "검출수"})
for c in ["Stonyhurst 경도(°)", "위도(°)", "Carrington 경도(°)", "Carrington 위도(°)"]:
    daily[c] = daily[c].round(3)

screen_daily = daily[daily["NOAA"].isin(SCREEN)].copy()

rot = rot.rename(columns={
    "ar_noaanum": "NOAA", "mean_lat": "평균 위도(°)", "n_days": "관측일수",
    "span_days": "추적기간(일)", "slope_degday": "Carr.경도 변화율(°/일)",
    "omega_sid_degday": "항성 회전각속도 Ω(°/일)", "fit_rms_deg": "피팅 RMS(°)",
    "on_screen": "화면표시"})
rot["화면표시"] = rot["화면표시"].map({True: "O", False: ""})
rot["회전주기(일)"] = (360.0 / rot["항성 회전각속도 Ω(°/일)"]).round(2)

# ---------- styling helpers ----------
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SCREEN_FILL = PatternFill("solid", fgColor="FCE4D6")
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F3864")
BODY = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center", vertical="center")

def write_df(ws, df, start_row=1, screen_col=None):
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(start_row, j, col)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CEN; c.border = BORDER
    for i, (_, row) in enumerate(df.iterrows(), start_row + 1):
        is_screen = screen_col and str(row.get(screen_col, "")) in ("O", "1")
        for j, col in enumerate(df.columns, 1):
            v = row[col]
            if isinstance(v, (np.integer,)): v = int(v)
            elif isinstance(v, (np.floating,)): v = float(v)
            c = ws.cell(i, j, v)
            c.font = BODY; c.alignment = CEN; c.border = BORDER
            if is_screen: c.fill = SCREEN_FILL
    for j, col in enumerate(df.columns, 1):
        w = max(len(str(col)), *(len(str(x)) for x in df[col])) if len(df) else len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 3, 11), 26)
    ws.freeze_panes = ws.cell(start_row + 1, 1)

wb = Workbook()

# ---- Sheet 1: 방법론 ----
ws = wb.active; ws.title = "방법론"
ws["A1"] = "태양 흑점 차등회전 분석 — 데이터 & 방법"
ws["A1"].font = TITLE_FONT
notes = [
    "",
    "■ 데이터 출처",
    "   NASA HEK (Heliophysics Event Knowledgebase) — Helioviewer 마커의 원천 데이터.",
    "   검출 소스: HMI SHARP (SDO/HMI) + NOAA SWPC Observer.",
    "   API: https://www.lmsal.com/hek/her  (event_type=AR)",
    "",
    "■ 관측 기간",
    "   2026-05-29 기준 ±2개월  →  2026-03-29 ~ 2026-07-29 (UTC).",
    "",
    "■ 좌표계",
    "   위도(°): 태양 적도 기준 heliographic latitude (Stonyhurst=Carrington 동일).",
    "   Stonyhurst 경도(°): 중앙자오선 기준 경도(-는 동쪽, +는 서쪽 림).",
    "   Carrington 경도(°): 태양과 함께 항성회전(25.38일, 14.1844°/일)하는 고정 좌표계.",
    "",
    "■ 회전 각속도 산출 방법",
    "   Carrington 경도는 태양이 항성 Carrington 속도(Ω_C=14.1844°/일)로 회전.",
    "   위도 φ에서 국소 회전율 Ω(φ)로 도는 흑점의 Carrington 경도는",
    "        dL_C/dt = Ω(φ) − Ω_C   로 표류.",
    "   따라서 AR별로 L_C(t)를 선형피팅한 기울기로",
    "        Ω_sid(φ) = 14.1844 + (기울기)   [°/일, 항성 기준]",
    "",
    "■ 차등회전 프로파일",
    "   Ω(φ) = A + B·sin²φ (+ C·sin⁴φ) 로 최소제곱 피팅(추적 신뢰도 가중).",
    "   본 데이터:  Ω = 14.50 − 4.81·sin²φ   (2항, N=40)",
    "   문헌값(Snodgrass & Ulrich 1990):  Ω = 14.71 − 2.39·sin²φ − 1.78·sin⁴φ",
    "",
    "■ 품질 필터 (회전율 표 & 그래프)",
    "   피팅 RMS ≤ 1.0°, 관측일수 ≥ 5, 추적기간 ≥ 5일, Ω∈[12,16]°/일 만 사용.",
    "   림 근처 투영오차 제거를 위해 |Stonyhurst 경도| ≤ 60° 만 집계.",
    "",
    "■ 시트 구성",
    "   · 화면 흑점 일별위치 : Helioviewer 화면의 활성영역 11개 일별 위도/경도",
    "   · 전체 AR 일별위치   : 기간 내 전체 활성영역 일별 위도/경도",
    "   · 위도별 회전각속도   : AR별 평균위도·회전각속도·회전주기 (주황=화면 흑점)",
]
for k, line in enumerate(notes, 2):
    c = ws.cell(k, 1, line); c.font = Font(name="Arial", size=11,
        bold=line.startswith("■"), color="C0392B" if line.startswith("■") else "000000")
ws.column_dimensions["A"].width = 95

# ---- Sheet 2: 화면 흑점 일별위치 ----
ws2 = wb.create_sheet("화면 흑점 일별위치")
write_df(ws2, screen_daily)

# ---- Sheet 3: 전체 AR 일별위치 ----
ws3 = wb.create_sheet("전체 AR 일별위치")
write_df(ws3, daily)

# ---- Sheet 4: 위도별 회전각속도 ----
ws4 = wb.create_sheet("위도별 회전각속도")
rot_sorted = rot.sort_values("평균 위도(°)")
write_df(ws4, rot_sorted, screen_col="화면표시")

wb.save("흑점_차등회전_데이터.xlsx")
print("saved 흑점_차등회전_데이터.xlsx")
print("screen_daily rows:", len(screen_daily), "| all daily:", len(daily),
      "| rotation rows:", len(rot_sorted))
